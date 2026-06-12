import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ingest import load_mapping_table, DISCIPLINE_MAPPING

def run_qa(prepared_file_path, db_export_csv_path, output_path="QA_Report.xlsx"):
    print(f"Loading Mapping Table...")
    mapping_dict = load_mapping_table()
    
    print(f"Loading Prepared Data: {prepared_file_path}")
    df_prep = pd.read_excel(prepared_file_path)
    
    # Identify columns in prepared data
    avail_id_col = next((c for c in df_prep.columns if 'availability id' in str(c).lower()), None)
    int_id_col = next((c for c in df_prep.columns if str(c).strip().lower() == '_id'), None)
    if not int_id_col:
        int_id_col = next((c for c in df_prep.columns if str(c).strip().lower() == 'id'), None)
    fac_col = next((c for c in df_prep.columns if 'facility' in str(c).lower() or 'setting' in str(c).lower()), None)
    disc_col = next((c for c in df_prep.columns if 'discipline' in str(c).lower()), None)
    name_col = next((c for c in df_prep.columns if 'name' in str(c).lower() and 'discipline' not in str(c).lower()), None)

    if not all([avail_id_col, fac_col, disc_col, name_col]):
        print(f"Error: Prepared data is missing one or more required columns.")
        print(f"Found: ID={avail_id_col}, Fac={fac_col}, Disc={disc_col}, Name={name_col}")
        return False

    print(f"Loading DB Export Data: {db_export_csv_path}")
    df_db = pd.read_csv(db_export_csv_path)
    
    # Identify curriculum columns dynamically
    curriculum_cols = []
    for c in df_db.columns:
        if 'curriculum' in c and 'disciplineId' in c:
            prefix = c.split('.disciplineId')[0] # e.g. "curriculum[0]"
            spec_col = f"{prefix}.specializationId"
            if spec_col in df_db.columns:
                curriculum_cols.append((c, spec_col))

    results = []

    for index, row in df_prep.iterrows():
        avail_id = str(row[avail_id_col]).strip()
        # Handle floats like 10000001416.0
        if avail_id.endswith('.0'):
            avail_id = avail_id[:-2]
            
        avail_name = str(row[name_col]).strip() if name_col else ""
        fac_type = str(row[fac_col]).strip().upper()
        raw_disciplines = str(row[disc_col])

        # 1. Determine Expected Pairs
        expected_pairs = []
        if pd.notna(raw_disciplines) and raw_disciplines.lower() != 'nan':
            for d in raw_disciplines.split(','):
                d_clean = d.strip().lower()
                acronym = DISCIPLINE_MAPPING.get(d_clean)
                if acronym:
                    specs = mapping_dict.get((acronym, fac_type), [])
                    for s in specs:
                        # Convert UUIDs to lowercase for reliable comparison
                        expected_pairs.append((
                            str(s['disciplineId']).strip().lower(), 
                            str(s['specializationId']).strip().lower()
                        ))
        
        # Remove duplicates from expected_pairs
        expected_pairs = list(set(expected_pairs))

        # 2. Find corresponding DB Row
        db_row = None
        # Try match by displayId first (displayId == Availability ID)
        if 'displayId' in df_db.columns:
            db_matches = df_db[df_db['displayId'].astype(str).str.replace('.0', '', regex=False) == avail_id]
            if not db_matches.empty:
                db_row = db_matches.iloc[0]
        
        # Fallback: match by name
        if db_row is None and 'name' in df_db.columns:
            db_matches = df_db[df_db['name'].astype(str).str.strip() == avail_name]
            if not db_matches.empty:
                db_row = db_matches.iloc[0]
                
        # Extract DB Values
        db_display_id = ""
        db_id = ""
        db_name = ""
        actual_pairs = []
        
        if db_row is not None:
            if 'displayId' in df_db.columns:
                db_display_id = str(db_row['displayId'])
            if '_id' in df_db.columns:
                db_id = str(db_row['_id'])
            if 'name' in df_db.columns:
                db_name = str(db_row['name'])
                
            for disc_col_name, spec_col_name in curriculum_cols:
                val_disc = str(db_row[disc_col_name]).strip().lower()
                val_spec = str(db_row[spec_col_name]).strip().lower()
                if val_disc and val_disc != 'nan' and val_spec and val_spec != 'nan':
                    actual_pairs.append((val_disc, val_spec))

        # 3. Grade Pass/Fail
        status = "FAIL"
        reason = ""
        if db_row is None:
            reason = "Record not found in DB Export"
        elif not expected_pairs:
            reason = "No expected specializations mapped"
        else:
            # Check if all expected pairs exist in actual pairs
            missing = [p for p in expected_pairs if p not in actual_pairs]
            if not missing:
                status = "PASS"
            else:
                reason = f"Missing {len(missing)} expected combos"

        # Format pairs for output
        exp_str = "\\n".join([f"D: {d}\\nS: {s}" for d, s in expected_pairs])
        act_str = "\\n".join([f"D: {d}\\nS: {s}" for d, s in actual_pairs])

        results.append({
            'Prepared_Availability_ID': avail_id,
            'Prepared_Name': avail_name,
            'DB_Display_ID': db_display_id,
            'DB__id': db_id,
            'DB_Name': db_name,
            'Expected_Combos': exp_str,
            'Actual_Combos': act_str,
            'Status': status,
            'Reason': reason
        })

    # Create Excel Output
    columns = [
        'Prepared_Availability_ID', 'Prepared_Name', 'DB_Display_ID', 'DB__id', 
        'DB_Name', 'Expected_Combos', 'Actual_Combos', 'Status', 'Reason'
    ]
    df_out = pd.DataFrame(results, columns=columns)
    
    # Save unstyled excel first
    df_out.to_excel(output_path, index=False)
    
    # Apply Color Coding
    wb = load_workbook(output_path)
    ws = wb.active
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Assuming Status is the 8th column (H)
    # 1: Prepared_Availability_ID, 2: Prepared_Name, 3: DB_Display_ID, 4: DB__id, 5: DB_Name, 6: Expected_Combos, 7: Actual_Combos, 8: Status, 9: Reason
    status_col_idx = 8
    
    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col_idx).value
        fill_color = green_fill if status_val == "PASS" else red_fill
        
        # Color the entire row
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill_color

    wb.save(output_path)
    print(f"QA Completed. Generated Report: {output_path}")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python qa_engine.py <prepared_data.xlsx> <db_export.csv> [output.xlsx]")
        sys.exit(1)
        
    prep = sys.argv[1]
    db_exp = sys.argv[2]
    out = sys.argv[3] if len(sys.argv) > 3 else "QA_Report.xlsx"
    
    run_qa(prep, db_exp, out)
